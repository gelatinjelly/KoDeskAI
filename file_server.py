from fastapi import FastAPI, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import fitz
import docx
import openpyxl
import requests
import tempfile
import os
import json

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

OLLAMA_URL = "http://localhost:11434/api/generate"
MODEL_NAME = "desktop-v7"

# ── 텍스트 추출 ──────────────────────────

def extract_pdf(path):
    doc = fitz.open(path)
    return "".join(page.get_text() for page in doc).strip()

def extract_word(path):
    d = docx.Document(path)
    return "\n".join(p.text for p in d.paragraphs if p.text.strip())

def extract_excel(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        result.append(f"[시트: {sheet}]")
        for row in ws.iter_rows(values_only=True):
            row_data = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in row_data):
                result.append("\t".join(row_data))
    return "\n".join(result)

def extract_text(path, filename):
    ext = filename.lower().split(".")[-1]
    if ext == "pdf":
        return extract_pdf(path)
    elif ext in ["docx", "doc"]:
        return extract_word(path)
    elif ext in ["xlsx", "xls"]:
        return extract_excel(path)
    elif ext == "txt":
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    return ""

# ── 스트리밍 응답 ──────────────────────────

def stream_ollama(prompt):
    response = requests.post(OLLAMA_URL, json={
        "model": MODEL_NAME,
        "prompt": prompt,
        "stream": True
    }, stream=True)
    for line in response.iter_lines():
        if line:
            data = json.loads(line)
            token = data.get("response", "")
            if token:
                yield f"data: {token}\n\n"
            if data.get("done"):
                break

def ask_ollama(prompt):
    response = requests.post(OLLAMA_URL, json={
        "model": MODEL_NAME,
        "prompt": prompt,
        "stream": False
    })
    return response.json().get("response", "")

# ── 엔드포인트 ────────────────────────────────

@app.post("/analyze")
async def analyze_file(file: UploadFile = File(...)):
    suffix = "." + file.filename.split(".")[-1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        text = extract_text(tmp_path, file.filename)
        if not text:
            return {"error": "텍스트를 추출할 수 없습니다."}
        text = text[:2000]
        prompt = f"""### 지시: 당신은 데스크탑 AI 비서입니다. 다음 문서 내용을 핵심 위주로 요약해줘.
### 입력: {text}
### 답변:"""
        summary = ask_ollama(prompt)
        return {"filename": file.filename, "text_preview": text[:300], "summary": summary}
    finally:
        os.unlink(tmp_path)

@app.post("/ask-with-file-stream")
async def ask_with_file_stream(
    file: UploadFile = File(...),
    question: str = Form(default="이 문서의 핵심 내용을 요약해줘")
):
    suffix = "." + file.filename.split(".")[-1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        text = extract_text(tmp_path, file.filename)
        if not text:
            return {"error": "텍스트를 추출할 수 없습니다."}
        text = text[:2000]
        prompt = f"""### 지시: 당신은 데스크탑 AI 비서입니다. 다음 문서를 참고해서 질문에 답해줘.
### 입력: [문서 내용]
{text}

[질문]
{question}
### 답변:"""
        return StreamingResponse(
            stream_ollama(prompt),
            media_type="text/event-stream"
        )
    finally:
        os.unlink(tmp_path)

@app.post("/ask-with-file")
async def ask_with_file(
    file: UploadFile = File(...),
    question: str = Form(default="이 문서의 핵심 내용을 요약해줘")
):
    suffix = "." + file.filename.split(".")[-1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        text = extract_text(tmp_path, file.filename)
        if not text:
            return {"error": "텍스트를 추출할 수 없습니다."}
        text = text[:2000]
        prompt = f"""### 지시: 당신은 데스크탑 AI 비서입니다. 다음 문서를 참고해서 질문에 답해줘.
### 입력: [문서 내용]
{text}

[질문]
{question}
### 답변:"""
        answer = ask_ollama(prompt)
        return {"filename": file.filename, "question": question, "answer": answer}
    finally:
        os.unlink(tmp_path)

@app.get("/health")
def health():
    return {"status": "ok"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)